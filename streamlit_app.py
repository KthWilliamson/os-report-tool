import streamlit as st
import csv
import openpyxl
import io
from collections import defaultdict
from datetime import datetime
from openpyxl.cell.cell import MergedCell

st.set_page_config(page_title="Order & Scale | Report Generator", layout="centered")

st.title("📊 Weekly Financial Report Generator")
st.write("Upload your Workamajig CSV and your previous report to update history.")

# --- FILE UPLOADERS ---
csv_file = st.file_uploader("1. Drop Workamajig CSV here", type=['csv'])
prev_report = st.file_uploader("2. Drop Previous Report (or Template) here", type=['xlsx'])

if st.button("Process Report"):
    if csv_file and prev_report:
        # Load files into memory
        wb = openpyxl.load_workbook(prev_report, data_only=False)
        
        # --- LOGIC CONSTANTS ---
        OMIT_PREFIX = "Yes-"
        START_ROW_OV = 10
        PROJ_NAME_COL, POP_COL, CURR_LABOR_COL, PTD_LABOR_COL = 2, 3, 5, 6

        # 1. PROCESS TRANSACTIONS
        ws_trans = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.worksheets[0]
        trans_header_map = {str(cell.value).strip(): cell.column for cell in ws_trans[1] if cell.value}
        
        # Clear old transaction data
        if ws_trans.max_row > 1:
            for row in ws_trans.iter_rows(min_row=2, max_row=ws_trans.max_row):
                for cell in row:
                    if not isinstance(cell, MergedCell): cell.value = None

        current_period_totals = defaultdict(float)
        project_date_ranges = {}
        unique_projects = set()

        # Read CSV
        decoded_file = csv_file.getvalue().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)
        next_trans_row = 2
        
        for row in reader:
            # Populate Transactions Tab
            for col_name, col_idx in trans_header_map.items():
                if col_name in row:
                    target_cell = ws_trans.cell(row=next_trans_row, column=col_idx)
                    if not isinstance(target_cell, MergedCell):
                        val = row[col_name]
                        if col_name in ['Quantity', 'Net', 'Gross']:
                            try: val = float(val.replace(',', ''))
                            except: pass
                        target_cell.value = val
            
            # Logic for Overview
            full_name = row.get('Project Full Name', '').strip()
            if full_name and not full_name.startswith(OMIT_PREFIX):
                unique_projects.add(full_name)
                raw_date = row.get('Expense Date', '')
                if raw_date:
                    try:
                        current_date = datetime.strptime(raw_date, '%m/%d/%Y')
                        if full_name not in project_date_ranges:
                            project_date_ranges[full_name] = [current_date, current_date]
                        else:
                            project_date_ranges[full_name][0] = min(project_date_ranges[full_name][0], current_date)
                            project_date_ranges[full_name][1] = max(project_date_ranges[full_name][1], current_date)
                    except ValueError: pass

                if row.get('Tran Type', '').strip().upper() == 'LABOR':
                    try: current_period_totals[full_name] += float(row['Gross'].replace(',', ''))
                    except: pass
            next_trans_row += 1

        # 2. UPDATE OVERVIEW
        ws_ov = wb["Account Overview"] if "Account Overview" in wb.sheetnames else wb.worksheets[2]
        existing_rows = {}
        for r in range(START_ROW_OV, ws_ov.max_row + 1):
            name = ws_ov.cell(row=r, column=PROJ_NAME_COL).value
            if name: existing_rows[str(name).strip()] = r

        sorted_projects = sorted(list(unique_projects))
        current_ov_row = START_ROW_OV

        for proj in sorted_projects:
            target_row = existing_rows.get(proj, current_ov_row)
            if not proj in existing_rows and "Total" in str(ws_ov.cell(row=target_row, column=PROJ_NAME_COL).value or ""):
                ws_ov.insert_rows(target_row)

            ws_ov.cell(row=target_row, column=PROJ_NAME_COL).value = proj
            if proj in project_date_ranges:
                s, e = project_date_ranges[proj]
                ws_ov.cell(row=target_row, column=POP_COL).value = f"{s.strftime('%m/%d/%y')} - {e.strftime('%m/%d/%y')}"

            curr_val = current_period_totals.get(proj, 0)
            prev_ptd_val = ws_ov.cell(row=target_row, column=PTD_LABOR_COL).value or 0
            try:
                if isinstance(prev_ptd_val, str): prev_ptd_val = float(prev_ptd_val.replace('$', '').replace(',', ''))
            except: prev_ptd_val = 0
            
            new_ptd_val = float(prev_ptd_val) + float(curr_val)
            ws_ov.cell(row=target_row, column=CURR_LABOR_COL).value = curr_val
            ws_ov.cell(row=target_row, column=PTD_LABOR_COL).value = new_ptd_val
            
            ws_ov.cell(row=target_row, column=CURR_LABOR_COL).number_format = '"$"#,##0.00'
            ws_ov.cell(row=target_row, column=PTD_LABOR_COL).number_format = '"$"#,##0.00'

            if not proj in existing_rows: current_ov_row = max(current_ov_row, target_row) + 1

        # --- DOWNLOAD BUTTON ---
        output = io.BytesIO()
        wb.save(output)
        st.success("Processing Complete!")
        st.download_button(
            label="💾 Download Updated Report",
            data=output.getvalue(),
            file_name="Updated_Client_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Please upload both files to proceed.")